"""
REPORTE DE TRANSFERENCIAS BANCARIAS
Basado en el Motor V32 — Agrupa por Fondo (una pestaña por fondo)
Columnas: #, Certificado, Inversionista, Moneda, Banco, N° Cuenta, CCI, NETO FINAL
"""
import sys, os
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# Ajuste de ruta: El script ahora está en FLOW_CHARTS/scripts_CRM/
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.data.supabase_client import get_supabase_client
from FLOW_CHARTS.scripts_CRM.C01_Motores_Calculo_NAV_y_PYL.CALCULO_Retornos_Intereses_V32 import generate_retornos_v32

PRIORITY_FONDOS = ["NSGPEN01", "NSGPEN02", "NSGPEN03", "NSGUSD01", "NSGUSD02", "NSLCON01"]

def run():
    print("🚀 Corriendo Motor V32 para generar reporte de transferencias...")
    fondos_dict, _ = generate_retornos_v32(return_data=True)

    print("📋 Cargando datos de inversionistas desde CRM...")
    sb = get_supabase_client()

    # Cargar todos los inversionistas en un mapa por documento_identidad y codigo_inversionista
    inv_res = sb.table('crm_inversionistas').select(
        'documento_identidad, codigo_inversionista, nombre_completo, '
        'banco_nombre_pen, numero_cuenta_pen, cci_pen, '
        'banco_nombre_usd, numero_cuenta_usd, cci_usd'
    ).execute()

    inv_map = {}
    for i in inv_res.data:
        for key in ['documento_identidad', 'codigo_inversionista']:
            val = i.get(key)
            if val:
                inv_map[str(val).strip().lower()] = i

    # Cargar certificados para obtener id_inversionista_1 y moneda
    cert_res = sb.table('crm_certificados').select(
        'id_certificado, crm_contratos(id_fondo, moneda, id_inversionista_1)'
    ).eq('estado', 'emitido').execute()

    cert_meta = {}
    for c in cert_res.data:
        contrato = c.get('crm_contratos') or {}
        cert_meta[c['id_certificado']] = {
            'moneda':        contrato.get('moneda', 'PEN'),
            'id_inv_1':      str(contrato.get('id_inversionista_1', '') or '').strip().lower(),
            'id_fondo':      contrato.get('id_fondo', ''),
        }

    # Construir el reporte agrupado por fondo
    output_dir = project_root / "reports"
    output_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = output_dir / f"transferencias_bancarias_v32_{ts}.xlsx"

    # Ordenar fondos con prioridad
    fondos_ordenados = [f for f in PRIORITY_FONDOS if f in fondos_dict] + \
                       [f for f in fondos_dict if f not in PRIORITY_FONDOS]

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for fondo_id in fondos_ordenados:
            rows_excel = fondos_dict[fondo_id]
            filas_reporte = []
            contador = 1

            for row in rows_excel:
                cert_id = row.get("Certificado", "")

                # Omitir filas de aumentos (no tienen # asignado = filas hijas)
                if not row.get("#"):
                    continue

                neto_final = float(row.get("NETO FINAL", 0) or 0)

                # Omitir si no hay nada que transferir
                if neto_final <= 0:
                    continue

                meta = cert_meta.get(cert_id, {})
                moneda  = meta.get('moneda', 'PEN')
                id_inv  = meta.get('id_inv_1', '')

                # Buscar datos bancarios del inversionista
                inv = inv_map.get(id_inv, {})
                nombre_inv = inv.get('nombre_completo', row.get('Inversionista', 'N/A'))

                if moneda == 'USD':
                    banco   = inv.get('banco_nombre_usd', 'NO REGISTRADO')
                    cuenta  = inv.get('numero_cuenta_usd', 'NO REGISTRADO')
                    cci     = inv.get('cci_usd', 'NO REGISTRADO')
                else:
                    banco   = inv.get('banco_nombre_pen', 'NO REGISTRADO')
                    cuenta  = inv.get('numero_cuenta_pen', 'NO REGISTRADO')
                    cci     = inv.get('cci_pen', 'NO REGISTRADO')

                filas_reporte.append({
                    "#":             contador,
                    "CERTIFICADO":   cert_id,
                    "INVERSIONISTA": nombre_inv,
                    "MONEDA":        moneda,
                    "BANCO":         banco,
                    "N° CUENTA":     cuenta,
                    "CCI":           cci,
                    "NETO FINAL":    neto_final,
                })
                contador += 1

            if not filas_reporte:
                print(f"  ⚠️  {fondo_id}: sin transferencias (NETO FINAL = 0 en todos los certificados).")
                continue

            # Fila de total
            total = sum(f["NETO FINAL"] for f in filas_reporte)
            filas_reporte.append({
                "#": "", "CERTIFICADO": "", "INVERSIONISTA": "",
                "MONEDA": "", "BANCO": "", "N° CUENTA": "TOTAL FONDO",
                "CCI": "", "NETO FINAL": total,
            })

            df = pd.DataFrame(filas_reporte)
            sheet_name = f"Fondo_{fondo_id[:24]}"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # ── FORMATO ──────────────────────────────────────
            ws = writer.sheets[sheet_name]
            header_fill  = PatternFill("solid", fgColor="004d40")
            total_fill   = PatternFill("solid", fgColor="e3bf60")
            header_font  = Font(bold=True, color="FFFFFF", size=10)
            total_font   = Font(bold=True, size=10)
            border_thin  = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Anchos de columna
            col_widths = [4, 36, 28, 8, 18, 22, 26, 14]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Dar formato a header
            for cell in ws[1]:
                cell.fill       = header_fill
                cell.font       = header_font
                cell.alignment  = Alignment(horizontal='center', vertical='center')
                cell.border     = border_thin

            # Dar formato a filas de datos
            last_row = len(filas_reporte) + 1
            for row_idx in range(2, last_row + 1):
                is_total = (row_idx == last_row)
                for col_idx in range(1, 9):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border_thin
                    cell.alignment = Alignment(vertical='center')
                    if is_total:
                        cell.fill = total_fill
                        cell.font = total_font
                # NETO FINAL con formato numérico
                neto_cell = ws.cell(row=row_idx, column=8)
                neto_cell.number_format = '#,##0.00'
                neto_cell.alignment = Alignment(horizontal='right', vertical='center')

            ws.row_dimensions[1].height = 20
            ws.freeze_panes = "A2"

            print(f"  ✅ {fondo_id}: {len(filas_reporte)-1} transferencia(s) | Total: {total:,.2f}")

    print(f"\n📂 Reporte guardado en: {output_path}")
    return output_path

if __name__ == "__main__":
    run()
